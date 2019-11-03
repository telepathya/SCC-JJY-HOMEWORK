import requests as req
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from pymongo import MongoClient           # pymongo를 임포트 하기(패키지 인스톨 먼저 해야겠죠?)
client = MongoClient('localhost', 27017)  # mongoDB는 27017 포트로 돌아갑니다.
db = client.dbsparta

headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
data = req.get( "https://www.genie.co.kr/chart/top200?ditc=D&rtm=N&ymd=20190908", headers=headers )

soup = BeautifulSoup( data.text, 'html.parser' )

# print( soup )

charts = soup.select( "#body-content > div.newest-list > div > table > tbody > tr" )
#work_book = load_workbook('practice.xlsx')
#work_sheet = work_book['prac']



for i in range( 50 ) :
    rank = i+1
    title = charts[i].select_one("td.info > a.title.ellipsis").text.strip()
    singer = charts[i].select_one("td.info > a.artist.ellipsis").text.strip()

    #db.homework.insert_one({'rank':i+1})
    #db.homework.insert_one({'title':title})
    #db.homework.insert_one({'singer':singer})

    dic = {'rank':i+1,'title':title,'singer':singer}
    db.homework.insert_one(dic)

    print(dic)

    #work_sheet.cell(row = rank, column = 1, value = rank )
    #work_sheet.cell(row = rank, column = 2, value = title)
    #work_sheet.cell(row = rank, column = 3, value = singer)

    #print( rank, title, singer )

# i = 0
# for chart in charts :
#    i+=1
#     rank = i
#     title = charts.select_one("td.info > a.title.ellipsis").text.strip()
#     singer = charts.select_one("td.info > a.artist.ellipsis").text.strip()
#     work_sheet.cell(row = rank, column=1, value = rank )
#     work_sheet.cell(row = rank, column=2, value = title)
#     work_sheet.cell(row = rank, column=3, value = singer)
#     # print( rank, title, singer )
#
#work_book.save('practice.xlsx')
#
# print(charts)
