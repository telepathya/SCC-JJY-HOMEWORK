from openpyxl import load_workbook

work_book = load_workbook('practice.xlsx')
work_sheet = work_book['prac']

# 데이터를 읽어봅니다.
temp_text = work_sheet.cell(row = 1, column = 2).value

print(temp_text)

